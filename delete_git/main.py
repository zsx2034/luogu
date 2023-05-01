import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

data = pd.read_csv("data.csv",names=["item", "step", "machine", "start_time", "end_time", "time_len", "square_time"], header=None)

workbook = Workbook()
sheet = workbook.create_sheet("sheet1")

# store nine colors
colors = ['FF0000', '00FF00', '0000FF', 'FFFF00', '00FFFF', 'FF00FF', 'C0C0C0', '808080', '800000']
# transfer colors to PatternFill
fills = [PatternFill("solid", fgColor=x) for x in colors]
i = 1
for line in data.itertuples():
    machine = getattr(line, 'machine')
    start_time = getattr(line, 'start_time')
    end_time = getattr(line, 'end_time')
    item = getattr(line, 'item')

    for col in range(start_time, end_time):
        cell = sheet.cell(row=8 - machine + 2, column=col + 2)
        cell.value = str(item) + "  " + str(i)
        cell.fill = fills[item-1]
    i += 1
workbook.save("result.xlsx")
